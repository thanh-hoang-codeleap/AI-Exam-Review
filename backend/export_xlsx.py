import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font


def load_json_data(file_path):
    with open(file_path) as file:
        return json.load(file)


def flatten_data(data):
    flattened_data = []

    # Ensure 'output' is in data
    if "output" not in data:
        return flattened_data  # Return empty if no 'output' key

    for json_string in data["output"]:
        parsed_data = json_string

        # Extract original and corrected sentence with checks
        original = parsed_data.get("original", "")
        corrected = parsed_data.get("corrected", "")

        # Ensure 'mistakes' key exists
        if "mistakes" in parsed_data:
            for mistake in parsed_data["mistakes"]:
                flattened_data.append(
                    {
                        "Original Sentence": original,
                        "Corrected Sentence": corrected,
                        "Incorrect Word/ Phrase": mistake.get("incorrect", ""),
                        "Correction for Word/ Phrase": mistake.get("correction", ""),
                        "Mistake Short-form": mistake.get("short_form", ""),
                        "Mistake Analysis": mistake.get("analysis", ""),
                    }
                )

    return flattened_data


def save_data_to_excel(data, file_path):
    df = pd.DataFrame(data)
    df.to_excel(file_path, index=False)


def load_workbook_from_excel(file_path):
    return load_workbook(file_path)


def merge_cells(ws, column):
    prev_value = None
    start_row = None
    for row in range(2, ws.max_row + 1):  # Skip header row (row 1)
        cell_value = ws.cell(row=row, column=column).value
        if cell_value == prev_value:
            if start_row is None:
                start_row = row - 1
        else:
            if start_row is not None and start_row < row - 1:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=column,
                    end_row=row - 1,
                    end_column=column,
                )
            start_row = None
        prev_value = cell_value

    if start_row is not None and start_row < ws.max_row:
        ws.merge_cells(
            start_row=start_row,
            start_column=column,
            end_row=ws.max_row,
            end_column=column,
        )


def add_borders(ws):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    dashed_border = Border(
        left=Side(style="dashed"),
        right=Side(style="dashed"),
        top=Side(style="dashed"),
        bottom=Side(style="dashed"),
    )

    prev_value = None
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border  # Apply normal thin border to all cells

        if ws.cell(row=row, column=1).value == prev_value:
            ws.cell(
                row=row - 1, column=1
            ).border = (
                dashed_border  # Dashed border for the row before this merged group
            )

        prev_value = ws.cell(row=row, column=1).value


def format_column_3_based_on_5(ws):
    for row in range(2, ws.max_row + 1):  # Starting from the second row (skip header)
        column_5_value = ws.cell(row=row, column=5).value  # Get value of the 5th column

        if (
            column_5_value and "writing" in column_5_value.lower()
        ):  # Check if "writing" is in the 5th column
            ws.cell(row=row, column=3).font = Font(
                color="FF156082"
            )  # Blue font (Hex code for blue)
        else:
            ws.cell(row=row, column=3).font = Font(
                color="FF0000"
            )  # Red font (Hex code for red)


def process_excel(file_path):
    wb = load_workbook_from_excel(file_path)
    ws = wb.active

    # Merge the cells in 'original' (column 1) and 'corrected' (column 2)
    merge_cells(ws, 1)
    merge_cells(ws, 2)

    # Add borders to all cells, with dashed border for merged rows
    add_borders(ws)

    # Apply formatting to the 3rd column based on the 5th column
    format_column_3_based_on_5(ws)

    # Save the modified Excel file
    wb.save(file_path)


def create_excel(json_path, excel_path):
    # Load JSON data
    data = load_json_data(json_path)

    # Flatten the data
    flattened_data = flatten_data(data)

    # Save the flattened data to Excel
    save_data_to_excel(flattened_data, excel_path)

    # Process the Excel file with formatting
    process_excel(excel_path)
